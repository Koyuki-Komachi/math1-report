
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import japanize_matplotlib
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

def get_distinct_colors(n):
    """
    異なる色を生成する関数
    
    Args:
        n (int): 必要な色の数
    
    Returns:
        list: 異なる色のリスト
    """
    # Color Brewerのカラーパレット（質的カラーパレット）
    color_palette = [
        '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', 
        '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
        '#1a55FF', '#FF6B6B', '#4ECDC4', '#45B7D1', '#FDCB6E',
        '#6C5CE7', '#FF8A5B', '#2ECC71', '#3498DB', '#E74C3C'
    ]
    
    # 必要な色の数が既存のパレットを超える場合は、追加の色を生成
    if n > len(color_palette):
        # matplotlib の色マップから追加の色を取得
        additional_colors = plt.cm.Set3(np.linspace(0, 1, n - len(color_palette)))
        color_palette.extend([mcolors.rgb2hex(color) for color in additional_colors])
    
    return color_palette[:n]

def process_sdg_data():
    # ファイルパス
    input_file = 'SDR2024-data.xlsx'
    output_file = 'sdgs_subsahara_graph.xlsx'
    
    try:
        # データの読み込み
        raw_data_panel = pd.read_excel(input_file, sheet_name='Raw Data - Panel')
        backdated_sdg = pd.read_excel(input_file, sheet_name='Backdated SDG Index')
        
        # 'Country' と 'indexreg' の対応をマッピング
        country_to_indexreg = raw_data_panel.set_index('Country')['indexreg'].to_dict()
        
        # 'Backdated SDG Index' シートに 'indexreg' 列を追加
        backdated_sdg['indexreg'] = backdated_sdg['Country'].map(country_to_indexreg)
        
        # 年次とindexregごとの平均を計算
        years = list(range(2000, 2024))
        indexregs = backdated_sdg['indexreg'].dropna().unique()
        goals = [f'goal{i}' for i in range(1, 18)]
        
        averages_dict = {'year': []}
        for indexreg in indexregs:
            averages_dict[indexreg] = []
        
        for year in years:
            yearly_data = backdated_sdg[backdated_sdg['year'] == year]
            averages_dict['year'].append(year)
            for indexreg in indexregs:
                reg_data = yearly_data[yearly_data['indexreg'] == indexreg]
                if reg_data.empty:
                    averages_dict[indexreg].append(None)
                    continue
                # goal1 ~ goal17 の平均を計算（0とNaNを除外）
                goal_means = []
                for goal in goals:
                    valid_data = reg_data[goal].replace(0, pd.NA).dropna()
                    if not valid_data.empty:
                        goal_means.append(valid_data.mean())
                average = sum(goal_means) / len(goal_means) if goal_means else None
                averages_dict[indexreg].append(average)
        
        # 平均値のデータフレーム作成
        averages_df = pd.DataFrame(averages_dict)
        
        # 表をExcelに保存
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            averages_df.to_excel(writer, sheet_name='平均値', index=False)
        
        # グラフの作成
        plt.figure(figsize=(16, 10))
        
        # 異なる色とスタイルの生成
        colors = get_distinct_colors(len(indexregs))
        line_styles = ['-', '--', '-.', ':', '-', '--', '-.', ':']
        markers = ['o', 's', '^', 'D', 'v', 'p', '*', 'h']
        
        for i, indexreg in enumerate(indexregs):
            plt.plot(
                averages_df['year'], 
                averages_df[indexreg], 
                color=colors[i], 
                linestyle=line_styles[i % len(line_styles)],
                marker=markers[i % len(markers)],
                linewidth=2.5,
                markersize=8,
                label=indexreg
            )
        
        plt.title('各Indexregの年次SDG Index Scoreの平均値 (2000-2023)', fontsize=16)
        plt.xlabel('年', fontsize=14)
        plt.ylabel('SDG Index Scoreの平均値', fontsize=14)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=10)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.tight_layout()
        
        # グラフを画像として保存
        graph_image = 'sdg_averages_graph.png'
        plt.savefig(graph_image, dpi=300, bbox_inches='tight')
        plt.close()
        
        # Excelファイルにグラフを挿入
        wb = load_workbook(output_file)
        ws = wb.create_sheet(title='グラフ')
        
        img = OpenpyxlImage(graph_image)
        img.anchor = 'A1'
        ws.add_image(img)
        
        wb.save(output_file)
        print(f"処理が完了しました。結果は '{output_file}' に保存されました。")
    
    except FileNotFoundError:
        print(f"エラー: ファイル '{input_file}' が見つかりません。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    process_sdg_data()

